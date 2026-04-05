"""Format-specific streaming file processor implementations."""

import logging
import os
from typing import Any, Dict, Optional

from ..streaming import StreamingDataSource
from .base import ProcessingProgress, StreamingFileProcessor
from .sources import (
    StreamingCSVSource,
    StreamingExcelSource,
    StreamingJSONSource,
    StreamingNumbersSource,
    StreamingODSSource,
)

logger = logging.getLogger(__name__)

# Check for Excel streaming support
try:
    import openpyxl
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Check for ijson support
try:
    import ijson

    IJSON_AVAILABLE = True
except ImportError:
    IJSON_AVAILABLE = False


class StreamingExcelProcessor(StreamingFileProcessor):
    """Memory-efficient Excel file processor using openpyxl.iter_rows()."""

    def get_file_type(self) -> str:
        return "excel"

    def create_streaming_source(
        self, sheet_name: Optional[str] = None
    ) -> StreamingDataSource:
        """Create streaming source for Excel file."""
        return StreamingExcelSource(self.file_path, sheet_name, self.progress)

    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate Excel processing requirements."""
        try:
            if not OPENPYXL_AVAILABLE:
                raise ValueError(
                    "openpyxl library required for Excel streaming. "
                    "Install with: pip install openpyxl"
                )

            workbook = load_workbook(self.file_path, read_only=True, data_only=True)

            total_sheets = len(workbook.sheetnames)
            self.progress.total_sheets = total_sheets

            estimated_rows = 0
            estimated_memory_mb = 0

            if workbook.sheetnames:
                first_sheet = workbook[workbook.sheetnames[0]]
                max_row = first_sheet.max_row or 0
                max_col = first_sheet.max_column or 0
                estimated_rows = max_row * total_sheets
                estimated_memory_per_sheet = (max_row * max_col * 50) / (1024 * 1024)
                estimated_memory_mb = estimated_memory_per_sheet * total_sheets

            workbook.close()

            return {
                "estimated_rows": estimated_rows,
                "estimated_memory_mb": estimated_memory_mb,
                "total_sheets": total_sheets,
                "file_size_mb": os.path.getsize(self.file_path) / (1024 * 1024),
                "processing_approach": "streaming_with_openpyxl_iter_rows",
            }

        except Exception as e:
            logger.warning(f"Failed to estimate Excel requirements: {e}")
            return {
                "estimated_rows": 1000,
                "estimated_memory_mb": 10,
                "total_sheets": 1,
                "file_size_mb": os.path.getsize(self.file_path) / (1024 * 1024),
                "processing_approach": "fallback_estimation",
            }


class StreamingJSONProcessor(StreamingFileProcessor):
    """Memory-efficient JSON file processor using ijson streaming."""

    def get_file_type(self) -> str:
        return "json"

    def create_streaming_source(
        self, sheet_name: Optional[str] = None
    ) -> StreamingDataSource:
        """Create streaming source for JSON file."""
        return StreamingJSONSource(self.file_path, self.progress)

    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate JSON processing requirements."""
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
        estimated_memory_mb = file_size_mb * 3
        estimated_rows = int(file_size_mb * 100)

        return {
            "estimated_rows": estimated_rows,
            "estimated_memory_mb": estimated_memory_mb,
            "file_size_mb": file_size_mb,
            "processing_approach": "streaming_with_ijson"
            if IJSON_AVAILABLE
            else "chunked_fallback",
        }


class StreamingCSVProcessor(StreamingFileProcessor):
    """Enhanced CSV processor with adaptive chunk sizing."""

    def get_file_type(self) -> str:
        return "csv"

    def create_streaming_source(
        self, sheet_name: Optional[str] = None
    ) -> StreamingDataSource:
        """Create streaming source for CSV file."""
        return StreamingCSVSource(self.file_path, self.progress)

    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate CSV processing requirements."""
        try:
            with open(self.file_path, "r", encoding="utf-8") as f:
                estimated_rows = sum(1 for _ in f) - 1

            file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)

            return {
                "estimated_rows": estimated_rows,
                "estimated_memory_mb": file_size_mb * 1.5,
                "file_size_mb": file_size_mb,
                "processing_approach": "pandas_chunksize_streaming",
            }

        except Exception as e:
            logger.warning(f"Failed to estimate CSV requirements: {e}")
            file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
            return {
                "estimated_rows": int(file_size_mb * 200),
                "estimated_memory_mb": file_size_mb * 1.5,
                "file_size_mb": file_size_mb,
                "processing_approach": "fallback_estimation",
            }


class StreamingODSProcessor(StreamingFileProcessor):
    """Memory-efficient ODS file processor."""

    def get_file_type(self) -> str:
        return "ods"

    def create_streaming_source(
        self, sheet_name: Optional[str] = None
    ) -> StreamingDataSource:
        """Create streaming source for ODS file."""
        return StreamingODSSource(self.file_path, sheet_name, self.progress)

    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate ODS processing requirements."""
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)

        return {
            "estimated_rows": int(file_size_mb * 100),
            "estimated_memory_mb": file_size_mb * 4,
            "file_size_mb": file_size_mb,
            "processing_approach": "pandas_with_odf_engine",
        }


class StreamingNumbersProcessor(StreamingFileProcessor):
    """Memory-efficient Numbers file processor."""

    def get_file_type(self) -> str:
        return "numbers"

    def create_streaming_source(
        self, sheet_name: Optional[str] = None
    ) -> StreamingDataSource:
        """Create streaming source for Numbers file."""
        return StreamingNumbersSource(self.file_path, sheet_name, self.progress)

    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate Numbers processing requirements."""
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)

        return {
            "estimated_rows": int(file_size_mb * 80),
            "estimated_memory_mb": file_size_mb * 5,
            "file_size_mb": file_size_mb,
            "processing_approach": "numbers_parser_with_chunking",
        }
