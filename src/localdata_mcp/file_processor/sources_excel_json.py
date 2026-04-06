"""Streaming data sources for Excel and JSON formats."""

import gc
import logging
import os
from typing import Generator, Optional

import pandas as pd

from ..streaming import StreamingDataSource
from .base import ProcessingProgress

logger = logging.getLogger(__name__)

# Check for streaming JSON support
try:
    import ijson

    IJSON_AVAILABLE = True
except ImportError:
    IJSON_AVAILABLE = False

# Check for Excel streaming support
try:
    import openpyxl
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class StreamingExcelSource(StreamingDataSource):
    """True streaming Excel source using openpyxl.iter_rows()."""

    def __init__(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        progress: Optional[ProcessingProgress] = None,
    ):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.progress = progress
        self._workbook = None
        self._estimated_rows = None

    def get_chunk_iterator(
        self, chunk_size: int
    ) -> Generator[pd.DataFrame, None, None]:
        """Stream Excel data using openpyxl.iter_rows() for true memory efficiency."""
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl library required for Excel streaming")

        try:
            self._workbook = load_workbook(
                self.file_path, read_only=True, data_only=True
            )
            sheets_to_process = (
                [self.sheet_name] if self.sheet_name else self._workbook.sheetnames
            )

            for sheet_name in sheets_to_process:
                if self.progress:
                    self.progress.current_sheet = sheet_name
                    self.progress.sheets_processed += 1

                worksheet = self._workbook[sheet_name]

                if worksheet.max_row is None or worksheet.max_row == 0:
                    logger.warning(f"Excel sheet '{sheet_name}' is empty, skipping")
                    continue

                # Get headers from first row
                header_row = next(
                    worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
                )
                headers = [
                    str(cell) if cell is not None else f"Column_{i + 1}"
                    for i, cell in enumerate(header_row)
                ]
                headers = [
                    h.strip().replace(" ", "_").replace("-", "_") for h in headers
                ]

                # Stream data in chunks using iter_rows
                chunk_data = []
                row_count = 0

                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if all(cell is None for cell in row):
                        continue

                    row_data = [cell if cell is not None else "" for cell in row]

                    while len(row_data) < len(headers):
                        row_data.append("")
                    row_data = row_data[: len(headers)]

                    chunk_data.append(row_data)
                    row_count += 1

                    if len(chunk_data) >= chunk_size:
                        chunk_df = pd.DataFrame(chunk_data, columns=headers)
                        if self.progress:
                            self.progress.rows_processed += len(chunk_df)

                        yield chunk_df
                        chunk_data = []

                        if row_count % (chunk_size * 5) == 0:
                            gc.collect()

                if chunk_data:
                    chunk_df = pd.DataFrame(chunk_data, columns=headers)
                    if self.progress:
                        self.progress.rows_processed += len(chunk_df)
                    yield chunk_df

        except Exception as e:
            logger.error(f"Error streaming Excel file: {e}")
            raise
        finally:
            if self._workbook:
                self._workbook.close()

    def estimate_total_rows(self) -> Optional[int]:
        """Estimate total rows across all sheets."""
        if self._estimated_rows is not None:
            return self._estimated_rows

        try:
            workbook = load_workbook(self.file_path, read_only=True)
            total_rows = 0

            sheets_to_process = (
                [self.sheet_name] if self.sheet_name else workbook.sheetnames
            )
            for sheet_name in sheets_to_process:
                worksheet = workbook[sheet_name]
                if worksheet.max_row:
                    total_rows += max(0, worksheet.max_row - 1)

            workbook.close()
            self._estimated_rows = total_rows
            return total_rows
        except Exception:
            return None

    def estimate_memory_per_row(self) -> float:
        """Estimate memory usage per row."""
        return 200.0


class StreamingJSONSource(StreamingDataSource):
    """Streaming JSON source using ijson for memory efficiency."""

    def __init__(self, file_path: str, progress: Optional[ProcessingProgress] = None):
        self.file_path = file_path
        self.progress = progress
        self._estimated_rows = None

    def get_chunk_iterator(
        self, chunk_size: int
    ) -> Generator[pd.DataFrame, None, None]:
        """Stream JSON data using ijson for memory efficiency."""
        if IJSON_AVAILABLE:
            yield from self._stream_with_ijson(chunk_size)
        else:
            logger.warning(
                "ijson not available, using fallback chunked loading for JSON"
            )
            yield from self._stream_fallback(chunk_size)

    def _stream_with_ijson(
        self, chunk_size: int
    ) -> Generator[pd.DataFrame, None, None]:
        """Stream JSON using ijson library."""
        try:
            chunk_data = []

            with open(self.file_path, "rb") as f:
                parser = ijson.parse(f)
                current_item = {}
                in_array_item = False

                for prefix, event, value in parser:
                    if event == "start_map" and prefix.endswith(".item"):
                        in_array_item = True
                        current_item = {}
                    elif event == "end_map" and in_array_item:
                        if current_item:
                            chunk_data.append(current_item)
                            if self.progress:
                                self.progress.rows_processed += 1

                            if len(chunk_data) >= chunk_size:
                                chunk_df = pd.json_normalize(chunk_data)
                                yield chunk_df
                                chunk_data = []

                        in_array_item = False
                    elif in_array_item and event in (
                        "string",
                        "number",
                        "boolean",
                        "null",
                    ):
                        key = prefix.split(".")[-1]
                        current_item[key] = value

            if chunk_data:
                chunk_df = pd.json_normalize(chunk_data)
                yield chunk_df

        except Exception as e:
            logger.error(f"Error streaming JSON with ijson: {e}")
            yield from self._stream_fallback(chunk_size)

    def _stream_fallback(self, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
        """Fallback streaming by loading JSON and chunking."""
        try:
            df = pd.read_json(self.file_path)

            for start in range(0, len(df), chunk_size):
                chunk = df.iloc[start : start + chunk_size].copy()
                if not chunk.empty:
                    if self.progress:
                        self.progress.rows_processed += len(chunk)
                    yield chunk
        except Exception as e:
            logger.error(f"Error in JSON fallback streaming: {e}")
            raise

    def estimate_total_rows(self) -> Optional[int]:
        """Estimate total rows in JSON file."""
        if self._estimated_rows is not None:
            return self._estimated_rows

        try:
            if IJSON_AVAILABLE:
                with open(self.file_path, "rb") as f:
                    row_count = 0
                    for _ in ijson.items(f, "item"):
                        row_count += 1
                    self._estimated_rows = row_count
            else:
                file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
                self._estimated_rows = int(file_size_mb * 100)

            return self._estimated_rows
        except Exception:
            return None

    def estimate_memory_per_row(self) -> float:
        """Estimate memory usage per row."""
        return 300.0
