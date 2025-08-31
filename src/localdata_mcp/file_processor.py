"""Streaming File Processing Pipeline for LocalData MCP.

This module implements memory-efficient streaming file processors that replace
batch DataFrame loading with streaming approaches. Works with the streaming
executor architecture to provide memory-bounded file processing.
"""

import gc
import json
import logging
import os
import tempfile
import time
from abc import ABC, abstractmethod
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Generator, List, Optional, Tuple, Union

import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.engine import Engine

# Import streaming infrastructure
from .streaming_executor import StreamingDataSource, StreamingQueryExecutor, create_streaming_source

logger = logging.getLogger(__name__)

# Check for streaming JSON support
try:
    import ijson
    IJSON_AVAILABLE = True
except ImportError:
    IJSON_AVAILABLE = False

# Check for Excel streaming support  
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Check for ODS support
try:
    from odf import opendocument, table
    ODFPY_AVAILABLE = True
except ImportError:
    ODFPY_AVAILABLE = False

# Check for Numbers support
try:
    from numbers_parser import Document
    NUMBERS_PARSER_AVAILABLE = True
except ImportError:
    NUMBERS_PARSER_AVAILABLE = False


@dataclass
class ProcessingProgress:
    """Progress tracking for file processing operations."""
    file_path: str
    file_type: str
    total_sheets: Optional[int] = None
    sheets_processed: int = 0
    current_sheet: Optional[str] = None
    estimated_rows: Optional[int] = None
    rows_processed: int = 0
    memory_usage_mb: float = 0.0
    processing_time_seconds: float = 0.0
    start_time: float = 0.0


class StreamingFileProcessor(ABC):
    """Abstract base class for streaming file processors."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.progress = ProcessingProgress(
            file_path=file_path,
            file_type=self.get_file_type(),
            start_time=time.time()
        )
    
    @abstractmethod
    def get_file_type(self) -> str:
        """Get the file type this processor handles."""
        pass
    
    @abstractmethod
    def create_streaming_source(self, sheet_name: Optional[str] = None) -> StreamingDataSource:
        """Create streaming data source for this file."""
        pass
    
    @abstractmethod
    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate memory and processing requirements."""
        pass
    
    def get_progress(self) -> ProcessingProgress:
        """Get current processing progress."""
        self.progress.processing_time_seconds = time.time() - self.progress.start_time
        return self.progress


class StreamingExcelProcessor(StreamingFileProcessor):
    """Memory-efficient Excel file processor using openpyxl.iter_rows()."""
    
    def get_file_type(self) -> str:
        return "excel"
    
    def create_streaming_source(self, sheet_name: Optional[str] = None) -> StreamingDataSource:
        """Create streaming source for Excel file."""
        return StreamingExcelSource(self.file_path, sheet_name, self.progress)
    
    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate Excel processing requirements."""
        try:
            if not OPENPYXL_AVAILABLE:
                raise ValueError("openpyxl library required for Excel streaming. Install with: pip install openpyxl")
            
            workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            
            total_sheets = len(workbook.sheetnames)
            self.progress.total_sheets = total_sheets
            
            # Sample first sheet to estimate row size
            estimated_rows = 0
            estimated_memory_mb = 0
            
            if workbook.sheetnames:
                first_sheet = workbook[workbook.sheetnames[0]]
                
                # Count rows efficiently
                max_row = first_sheet.max_row or 0
                max_col = first_sheet.max_column or 0
                estimated_rows = max_row * total_sheets
                
                # Estimate memory: ~50 bytes per cell on average
                estimated_memory_per_sheet = (max_row * max_col * 50) / (1024 * 1024)  # MB
                estimated_memory_mb = estimated_memory_per_sheet * total_sheets
            
            workbook.close()
            
            return {
                "estimated_rows": estimated_rows,
                "estimated_memory_mb": estimated_memory_mb,
                "total_sheets": total_sheets,
                "file_size_mb": os.path.getsize(self.file_path) / (1024 * 1024),
                "processing_approach": "streaming_with_openpyxl_iter_rows"
            }
            
        except Exception as e:
            logger.warning(f"Failed to estimate Excel requirements: {e}")
            return {
                "estimated_rows": 1000,
                "estimated_memory_mb": 10,
                "total_sheets": 1,
                "file_size_mb": os.path.getsize(self.file_path) / (1024 * 1024),
                "processing_approach": "fallback_estimation"
            }


class StreamingJSONProcessor(StreamingFileProcessor):
    """Memory-efficient JSON file processor using ijson streaming."""
    
    def get_file_type(self) -> str:
        return "json"
    
    def create_streaming_source(self, sheet_name: Optional[str] = None) -> StreamingDataSource:
        """Create streaming source for JSON file."""
        return StreamingJSONSource(self.file_path, self.progress)
    
    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate JSON processing requirements."""
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
        
        # For JSON, estimate based on file size
        # Typical JSON-to-DataFrame conversion has ~3x memory overhead
        estimated_memory_mb = file_size_mb * 3
        estimated_rows = int(file_size_mb * 100)  # Rough estimate: 100 rows per MB
        
        return {
            "estimated_rows": estimated_rows,
            "estimated_memory_mb": estimated_memory_mb,
            "file_size_mb": file_size_mb,
            "processing_approach": "streaming_with_ijson" if IJSON_AVAILABLE else "chunked_fallback"
        }


class StreamingCSVProcessor(StreamingFileProcessor):
    """Enhanced CSV processor with adaptive chunk sizing."""
    
    def get_file_type(self) -> str:
        return "csv"
    
    def create_streaming_source(self, sheet_name: Optional[str] = None) -> StreamingDataSource:
        """Create streaming source for CSV file."""
        return StreamingCSVSource(self.file_path, self.progress)
    
    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate CSV processing requirements."""
        try:
            # Quick line count for row estimation
            with open(self.file_path, 'r', encoding='utf-8') as f:
                estimated_rows = sum(1 for _ in f) - 1  # Subtract header
            
            file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
            
            return {
                "estimated_rows": estimated_rows,
                "estimated_memory_mb": file_size_mb * 1.5,  # CSV overhead is typically low
                "file_size_mb": file_size_mb,
                "processing_approach": "pandas_chunksize_streaming"
            }
            
        except Exception as e:
            logger.warning(f"Failed to estimate CSV requirements: {e}")
            file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
            return {
                "estimated_rows": int(file_size_mb * 200),  # Fallback estimate
                "estimated_memory_mb": file_size_mb * 1.5,
                "file_size_mb": file_size_mb,
                "processing_approach": "fallback_estimation"
            }


class StreamingODSProcessor(StreamingFileProcessor):
    """Memory-efficient ODS file processor."""
    
    def get_file_type(self) -> str:
        return "ods"
    
    def create_streaming_source(self, sheet_name: Optional[str] = None) -> StreamingDataSource:
        """Create streaming source for ODS file."""
        return StreamingODSSource(self.file_path, sheet_name, self.progress)
    
    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate ODS processing requirements."""
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
        
        return {
            "estimated_rows": int(file_size_mb * 100),
            "estimated_memory_mb": file_size_mb * 4,  # ODS has higher overhead
            "file_size_mb": file_size_mb,
            "processing_approach": "pandas_with_odf_engine"
        }


class StreamingNumbersProcessor(StreamingFileProcessor):
    """Memory-efficient Numbers file processor."""
    
    def get_file_type(self) -> str:
        return "numbers"
    
    def create_streaming_source(self, sheet_name: Optional[str] = None) -> StreamingDataSource:
        """Create streaming source for Numbers file."""
        return StreamingNumbersSource(self.file_path, sheet_name, self.progress)
    
    def estimate_processing_requirements(self) -> Dict[str, Any]:
        """Estimate Numbers processing requirements."""
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
        
        return {
            "estimated_rows": int(file_size_mb * 80),
            "estimated_memory_mb": file_size_mb * 5,  # Numbers has high overhead
            "file_size_mb": file_size_mb,
            "processing_approach": "numbers_parser_with_chunking"
        }


# Enhanced Streaming Data Sources

class StreamingExcelSource(StreamingDataSource):
    """True streaming Excel source using openpyxl.iter_rows()."""
    
    def __init__(self, file_path: str, sheet_name: Optional[str] = None, 
                 progress: Optional[ProcessingProgress] = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.progress = progress
        self._workbook = None
        self._estimated_rows = None
    
    def get_chunk_iterator(self, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
        """Stream Excel data using openpyxl.iter_rows() for true memory efficiency."""
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl library required for Excel streaming")
        
        try:
            self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            sheets_to_process = [self.sheet_name] if self.sheet_name else self._workbook.sheetnames
            
            for sheet_name in sheets_to_process:
                if self.progress:
                    self.progress.current_sheet = sheet_name
                    self.progress.sheets_processed += 1
                
                worksheet = self._workbook[sheet_name]
                
                if worksheet.max_row is None or worksheet.max_row == 0:
                    logger.warning(f"Excel sheet '{sheet_name}' is empty, skipping")
                    continue
                
                # Get headers from first row
                header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [str(cell) if cell is not None else f'Column_{i+1}' 
                          for i, cell in enumerate(header_row)]
                
                # Clean up headers
                headers = [h.strip().replace(' ', '_').replace('-', '_') for h in headers]
                
                # Stream data in chunks using iter_rows
                chunk_data = []
                row_count = 0
                
                # Start from row 2 (skip header)
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if all(cell is None for cell in row):
                        continue  # Skip empty rows
                    
                    # Convert row to list, handling None values
                    row_data = [cell if cell is not None else '' for cell in row]
                    
                    # Ensure row has same length as headers
                    while len(row_data) < len(headers):
                        row_data.append('')
                    row_data = row_data[:len(headers)]  # Trim if too long
                    
                    chunk_data.append(row_data)
                    row_count += 1
                    
                    if len(chunk_data) >= chunk_size:
                        # Yield chunk
                        chunk_df = pd.DataFrame(chunk_data, columns=headers)
                        if self.progress:
                            self.progress.rows_processed += len(chunk_df)
                        
                        yield chunk_df
                        chunk_data = []
                        
                        # Force garbage collection periodically
                        if row_count % (chunk_size * 5) == 0:
                            gc.collect()
                
                # Yield remaining data
                if chunk_data:
                    chunk_df = pd.DataFrame(chunk_data, columns=headers)
                    if self.progress:
                        self.progress.rows_processed += len(chunk_df)
                    yield chunk_df
        
        except Exception as e:
            logger.error(f"Error streaming Excel file: {e}")
            raise
        finally:
            if self._workbook:
                self._workbook.close()
    
    def estimate_total_rows(self) -> Optional[int]:
        """Estimate total rows across all sheets."""
        if self._estimated_rows is not None:
            return self._estimated_rows
        
        try:
            workbook = load_workbook(self.file_path, read_only=True)
            total_rows = 0
            
            sheets_to_process = [self.sheet_name] if self.sheet_name else workbook.sheetnames
            for sheet_name in sheets_to_process:
                worksheet = workbook[sheet_name]
                if worksheet.max_row:
                    total_rows += max(0, worksheet.max_row - 1)  # Subtract header row
            
            workbook.close()
            self._estimated_rows = total_rows
            return total_rows
        except Exception:
            return None
    
    def estimate_memory_per_row(self) -> float:
        """Estimate memory usage per row."""
        return 200.0  # Excel rows tend to be larger due to formatting


class StreamingJSONSource(StreamingDataSource):
    """Streaming JSON source using ijson for memory efficiency."""
    
    def __init__(self, file_path: str, progress: Optional[ProcessingProgress] = None):
        self.file_path = file_path
        self.progress = progress
        self._estimated_rows = None
    
    def get_chunk_iterator(self, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
        """Stream JSON data using ijson for memory efficiency."""
        if IJSON_AVAILABLE:
            yield from self._stream_with_ijson(chunk_size)
        else:
            # Fallback to chunked pandas loading
            logger.warning("ijson not available, using fallback chunked loading for JSON")
            yield from self._stream_fallback(chunk_size)
    
    def _stream_with_ijson(self, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
        """Stream JSON using ijson library."""
        try:
            chunk_data = []
            
            with open(self.file_path, 'rb') as f:
                # Parse JSON items stream
                parser = ijson.parse(f)
                current_item = {}
                in_array_item = False
                
                for prefix, event, value in parser:
                    if event == 'start_map' and prefix.endswith('.item'):
                        in_array_item = True
                        current_item = {}
                    elif event == 'end_map' and in_array_item:
                        if current_item:
                            chunk_data.append(current_item)
                            if self.progress:
                                self.progress.rows_processed += 1
                            
                            if len(chunk_data) >= chunk_size:
                                chunk_df = pd.json_normalize(chunk_data)
                                yield chunk_df
                                chunk_data = []
                        
                        in_array_item = False
                    elif in_array_item and event in ('string', 'number', 'boolean', 'null'):
                        key = prefix.split('.')[-1]
                        current_item[key] = value
                
                # Yield remaining data
                if chunk_data:
                    chunk_df = pd.json_normalize(chunk_data)
                    yield chunk_df
                    
        except Exception as e:
            logger.error(f"Error streaming JSON with ijson: {e}")
            # Fall back to regular loading and chunking
            yield from self._stream_fallback(chunk_size)
    
    def _stream_fallback(self, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
        """Fallback streaming by loading JSON and chunking."""
        try:
            df = pd.read_json(self.file_path)
            
            for start in range(0, len(df), chunk_size):
                chunk = df.iloc[start:start + chunk_size].copy()
                if not chunk.empty:
                    if self.progress:
                        self.progress.rows_processed += len(chunk)
                    yield chunk
        except Exception as e:
            logger.error(f"Error in JSON fallback streaming: {e}")
            raise
    
    def estimate_total_rows(self) -> Optional[int]:
        """Estimate total rows in JSON file."""
        if self._estimated_rows is not None:
            return self._estimated_rows
        
        try:
            if IJSON_AVAILABLE:
                # Count items using ijson
                with open(self.file_path, 'rb') as f:
                    row_count = 0
                    for _ in ijson.items(f, 'item'):
                        row_count += 1
                    self._estimated_rows = row_count
            else:
                # Quick estimate based on file size
                file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
                self._estimated_rows = int(file_size_mb * 100)  # Rough estimate
            
            return self._estimated_rows
        except Exception:
            return None
    
    def estimate_memory_per_row(self) -> float:
        """Estimate memory usage per row."""
        return 300.0  # JSON can have higher memory overhead


class StreamingCSVSource(StreamingDataSource):
    """Enhanced CSV streaming source with adaptive chunk sizing."""
    
    def __init__(self, file_path: str, progress: Optional[ProcessingProgress] = None, **kwargs):
        self.file_path = file_path
        self.progress = progress
        self.kwargs = kwargs
        self._estimated_rows = None
    
    def get_chunk_iterator(self, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
        """Stream CSV with enhanced error handling and adaptive chunking."""
        try:
            chunk_reader = pd.read_csv(self.file_path, chunksize=chunk_size, **self.kwargs)
            
            for chunk in chunk_reader:
                if not chunk.empty:
                    # Clean column names
                    chunk.columns = [str(col).strip().replace(' ', '_').replace('-', '_') 
                                   for col in chunk.columns]
                    
                    if self.progress:
                        self.progress.rows_processed += len(chunk)
                    
                    yield chunk
                    
        except pd.errors.ParserError:
            # Fallback for CSV with no header
            logger.warning("CSV parser error, trying without header")
            chunk_reader = pd.read_csv(self.file_path, chunksize=chunk_size, 
                                     header=None, **self.kwargs)
            for chunk in chunk_reader:
                if not chunk.empty:
                    if self.progress:
                        self.progress.rows_processed += len(chunk)
                    yield chunk
    
    def estimate_total_rows(self) -> Optional[int]:
        """Estimate total rows by counting lines."""
        if self._estimated_rows is not None:
            return self._estimated_rows
        
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                self._estimated_rows = sum(1 for _ in f) - 1  # Subtract header
            return self._estimated_rows
        except Exception:
            return None
    
    def estimate_memory_per_row(self) -> float:
        """Estimate memory usage per row."""
        return 100.0  # CSV typically has lower memory overhead


class StreamingODSSource(StreamingDataSource):
    """Streaming ODS source using pandas with odf engine."""
    
    def __init__(self, file_path: str, sheet_name: Optional[str] = None, 
                 progress: Optional[ProcessingProgress] = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.progress = progress
        self._sheets_data = None
    
    def get_chunk_iterator(self, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
        """Stream ODS data by loading sheets and chunking."""
        if not ODFPY_AVAILABLE:
            raise ValueError("odfpy library required for ODS files")
        
        try:
            # Load all sheets (we'll need to optimize this for very large files)
            with pd.ExcelFile(self.file_path, engine='odf') as excel_file:
                sheets_to_process = [self.sheet_name] if self.sheet_name else excel_file.sheet_names
                
                for sheet_name in sheets_to_process:
                    if self.progress:
                        self.progress.current_sheet = sheet_name
                        self.progress.sheets_processed += 1
                    
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='odf')
                    
                    if df.empty:
                        continue
                    
                    # Clean column names
                    df.columns = [str(col).strip().replace(' ', '_').replace('-', '_') 
                                for col in df.columns]
                    
                    # Yield in chunks
                    for start in range(0, len(df), chunk_size):
                        chunk = df.iloc[start:start + chunk_size].copy()
                        if not chunk.empty:
                            if self.progress:
                                self.progress.rows_processed += len(chunk)
                            yield chunk
        
        except Exception as e:
            logger.error(f"Error streaming ODS file: {e}")
            raise
    
    def estimate_total_rows(self) -> Optional[int]:
        """Estimate total rows across all sheets."""
        try:
            with pd.ExcelFile(self.file_path, engine='odf') as excel_file:
                total_rows = 0
                sheets_to_process = [self.sheet_name] if self.sheet_name else excel_file.sheet_names
                
                for sheet_name in sheets_to_process:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='odf')
                    total_rows += len(df)
                
                return total_rows
        except Exception:
            return None
    
    def estimate_memory_per_row(self) -> float:
        """Estimate memory usage per row."""
        return 200.0


class StreamingNumbersSource(StreamingDataSource):
    """Streaming Numbers source using numbers-parser."""
    
    def __init__(self, file_path: str, sheet_name: Optional[str] = None, 
                 progress: Optional[ProcessingProgress] = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.progress = progress
    
    def get_chunk_iterator(self, chunk_size: int) -> Generator[pd.DataFrame, None, None]:
        """Stream Numbers data by processing tables in chunks."""
        if not NUMBERS_PARSER_AVAILABLE:
            raise ValueError("numbers-parser library required for Numbers files")
        
        try:
            doc = Document(self.file_path)
            sheets_to_process = [s for s in doc.sheets if s.name == self.sheet_name] if self.sheet_name else doc.sheets
            
            for sheet in sheets_to_process:
                if self.progress:
                    self.progress.current_sheet = sheet.name
                    self.progress.sheets_processed += 1
                
                for table_idx, table in enumerate(sheet.tables):
                    table_data = table.rows(values_only=True)
                    
                    if not table_data or len(table_data) < 2:
                        continue
                    
                    headers = [str(h) if h is not None else f'Column_{i+1}' 
                             for i, h in enumerate(table_data[0])]
                    headers = [h.strip().replace(' ', '_').replace('-', '_').replace('.', '_') 
                             for h in headers]
                    
                    data_rows = table_data[1:]
                    
                    # Process in chunks
                    for start in range(0, len(data_rows), chunk_size):
                        chunk_rows = data_rows[start:start + chunk_size]
                        chunk_df = pd.DataFrame(chunk_rows, columns=headers)
                        
                        if not chunk_df.empty:
                            if self.progress:
                                self.progress.rows_processed += len(chunk_df)
                            yield chunk_df
        
        except Exception as e:
            logger.error(f"Error streaming Numbers file: {e}")
            raise
    
    def estimate_total_rows(self) -> Optional[int]:
        """Estimate total rows across all tables."""
        try:
            doc = Document(self.file_path)
            total_rows = 0
            
            sheets_to_process = [s for s in doc.sheets if s.name == self.sheet_name] if self.sheet_name else doc.sheets
            
            for sheet in sheets_to_process:
                for table in sheet.tables:
                    table_data = table.rows(values_only=True)
                    if table_data and len(table_data) > 1:
                        total_rows += len(table_data) - 1  # Subtract header
            
            return total_rows
        except Exception:
            return None
    
    def estimate_memory_per_row(self) -> float:
        """Estimate memory usage per row."""
        return 250.0


class FileProcessorFactory:
    """Factory for creating streaming file processors."""
    
    _processors = {
        'excel': StreamingExcelProcessor,
        'xlsx': StreamingExcelProcessor,
        'xlsm': StreamingExcelProcessor,
        'xls': StreamingExcelProcessor,
        'json': StreamingJSONProcessor,
        'csv': StreamingCSVProcessor,
        'tsv': StreamingCSVProcessor,
        'ods': StreamingODSProcessor,
        'numbers': StreamingNumbersProcessor,
    }
    
    @classmethod
    def create_processor(cls, file_path: str, file_type: str) -> StreamingFileProcessor:
        """Create appropriate streaming processor for file type.
        
        Args:
            file_path: Path to the file
            file_type: Type of file to process
            
        Returns:
            StreamingFileProcessor: Appropriate processor for the file type
            
        Raises:
            ValueError: If file type is not supported
        """
        if file_type not in cls._processors:
            raise ValueError(f"Unsupported file type for streaming: {file_type}")
        
        processor_class = cls._processors[file_type]
        return processor_class(file_path)
    
    @classmethod
    def get_supported_formats(cls) -> List[str]:
        """Get list of supported file formats."""
        return list(cls._processors.keys())
    
    @classmethod
    def is_supported(cls, file_type: str) -> bool:
        """Check if file type is supported for streaming processing."""
        return file_type in cls._processors


def create_streaming_file_engine(file_path: str, file_type: str, 
                                sheet_name: Optional[str] = None,
                                temp_files_registry: Optional[List[str]] = None) -> Tuple[Engine, Dict[str, Any]]:
    """Create SQLite engine from file using streaming processing.
    
    This function replaces the batch loading approach with memory-efficient
    streaming that processes files in chunks and uses temporary SQLite storage.
    
    Args:
        file_path: Path to the file to process
        file_type: Type of file (excel, csv, json, etc.)
        sheet_name: Specific sheet/table to load (if applicable)
        temp_files_registry: Optional list to track temporary files
        
    Returns:
        Tuple of (SQLite engine, processing_metadata)
    """
    start_time = time.time()
    
    # Create streaming processor
    processor = FileProcessorFactory.create_processor(file_path, file_type)
    
    # Get processing requirements estimate
    requirements = processor.estimate_processing_requirements()
    logger.info(f"Processing {file_type} file '{file_path}': {requirements}")
    
    # Determine if we need temporary file storage
    file_size_mb = requirements.get('file_size_mb', 0)
    use_temp_file = file_size_mb > 100  # Use temp file for files > 100MB
    
    # Create SQLite engine
    if use_temp_file:
        temp_fd, temp_path = tempfile.mkstemp(suffix=".sqlite", prefix="streaming_file_")
        os.close(temp_fd)
        
        if temp_files_registry is not None:
            temp_files_registry.append(temp_path)
        
        engine = create_engine(f"sqlite:///{temp_path}")
        logger.info(f"Using temporary SQLite file: {temp_path}")
    else:
        engine = create_engine("sqlite:///:memory:")
        logger.info("Using in-memory SQLite database")
    
    # Create streaming source
    streaming_source = processor.create_streaming_source(sheet_name)
    
    # Initialize streaming executor
    streaming_executor = StreamingQueryExecutor()
    
    # Process file in streaming chunks
    tables_created = {}
    sheet_count = 0
    total_rows_processed = 0
    
    try:
        # Create a temporary table name tracker
        used_table_names = set()
        
        chunk_size = 1000  # Start with reasonable chunk size
        chunk_number = 0
        
        for chunk_df in streaming_source.get_chunk_iterator(chunk_size):
            chunk_number += 1
            
            if chunk_df.empty:
                continue
            
            # Determine table name
            current_sheet = getattr(processor.progress, 'current_sheet', None)
            if current_sheet:
                table_name = _sanitize_table_name(current_sheet, used_table_names)
            else:
                table_name = 'data_table'
                used_table_names.add(table_name)
            
            # Handle multi-sheet files
            if table_name not in tables_created:
                # First chunk for this table - create table
                chunk_df.to_sql(table_name, engine, index=False, if_exists='replace')
                tables_created[table_name] = len(chunk_df)
                logger.info(f"Created table '{table_name}' with initial {len(chunk_df)} rows")
            else:
                # Append to existing table
                chunk_df.to_sql(table_name, engine, index=False, if_exists='append')
                tables_created[table_name] += len(chunk_df)
            
            total_rows_processed += len(chunk_df)
            
            # Progress logging
            if chunk_number % 10 == 0:
                logger.info(f"Processed {chunk_number} chunks, {total_rows_processed} total rows")
            
            # Adaptive chunk sizing based on performance
            if chunk_number > 1 and chunk_number % 5 == 0:
                # Simple adaptation: if processing is slow, reduce chunk size
                processing_time = time.time() - start_time
                if processing_time / chunk_number > 2.0:  # More than 2 seconds per chunk
                    chunk_size = max(chunk_size // 2, 100)
                    logger.info(f"Reduced chunk size to {chunk_size} due to slow processing")
    
    except Exception as e:
        logger.error(f"Error during streaming file processing: {e}")
        raise
    
    # Processing completed
    processing_time = time.time() - start_time
    progress = processor.get_progress()
    
    processing_metadata = {
        "file_path": file_path,
        "file_type": file_type,
        "processing_time_seconds": processing_time,
        "total_rows_processed": total_rows_processed,
        "tables_created": tables_created,
        "sheets_processed": progress.sheets_processed,
        "memory_approach": "streaming_chunks",
        "engine_type": "temporary_sqlite" if use_temp_file else "memory_sqlite",
        "requirements_estimate": requirements
    }
    
    logger.info(f"Streaming file processing completed: {total_rows_processed} rows in {processing_time:.3f}s, "
               f"{len(tables_created)} tables created")
    
    return engine, processing_metadata


def _sanitize_table_name(name: str, used_names: set) -> str:
    """Sanitize and ensure unique table name."""
    import re
    
    # Convert to string and strip whitespace
    name = str(name).strip()
    
    # Replace spaces and hyphens with underscores
    name = re.sub(r'[\s\-]+', '_', name)
    
    # Remove problematic characters
    name = re.sub(r'[^\w]', '_', name)
    
    # Remove consecutive underscores
    name = re.sub(r'_+', '_', name)
    
    # Ensure it starts with letter or underscore
    if name and not re.match(r'^[a-zA-Z_]', name):
        name = 'sheet_' + name
    
    # Handle empty names
    if not name:
        name = 'sheet_unnamed'
    
    # Ensure uniqueness
    original_name = name
    counter = 1
    while name.lower() in {n.lower() for n in used_names}:
        name = f"{original_name}_{counter}"
        counter += 1
    
    used_names.add(name)
    return name